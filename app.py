import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from PIL import Image, ImageDraw, ImageFont
import io
import time
import os
import tempfile

# ==========================================
# 設定・定数
# ==========================================
APP_TITLE = "Wireframe to Excel Specification Generator"
SHEET1_NAME = "原稿入力シート"
SHEET2_NAME = "ワイヤー確認用"

# 日本語フォントパス（環境に合わせて自動検出）
import platform

def get_japanese_font_path():
    """環境に応じた日本語フォントパスを返す"""
    system = platform.system()
    
    if system == "Darwin":  # macOS
        # macOSの日本語フォント候補
        mac_fonts = [
            "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
            "/System/Library/Fonts/Hiragino Sans GB.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/AppleSDGothicNeo.ttc",
        ]
        for font_path in mac_fonts:
            if os.path.exists(font_path):
                return font_path
    elif system == "Windows":
        # Windowsの日本語フォント候補
        win_fonts = [
            "C:/Windows/Fonts/meiryo.ttc",
            "C:/Windows/Fonts/msgothic.ttc",
            "C:/Windows/Fonts/YuGothM.ttc",
        ]
        for font_path in win_fonts:
            if os.path.exists(font_path):
                return font_path
    else:  # Linux
        linux_fonts = [
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK.ttc",
            "/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf",
            "/usr/share/fonts/truetype/takao-gothic/TakaoGothic.ttf",
        ]
        for font_path in linux_fonts:
            if os.path.exists(font_path):
                return font_path
    
    return None  # 見つからない場合はNone

FONT_PATH = get_japanese_font_path()

def setup_driver():
    """Headless Chromeの設定"""
    chrome_options = Options()
    # 新しいヘッドレスモードを使用（安定性向上）
    chrome_options.add_argument("--headless=new") 
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1280,800") # 初期ウィンドウサイズ
    
    # Streamlit Cloud（Linux）の場合はChromiumのパスを指定
    if os.path.exists("/usr/bin/chromium"):
        chrome_options.binary_location = "/usr/bin/chromium"
    elif os.path.exists("/usr/bin/chromium-browser"):
        chrome_options.binary_location = "/usr/bin/chromium-browser"
    
    try:
        # webdriver-managerを使用してChromeDriverを自動管理
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from webdriver_manager.core.os_manager import ChromeType
        
        # Chromiumを使う場合
        if chrome_options.binary_location:
            service = Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install())
        else:
            service = Service(ChromeDriverManager().install())
        
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except Exception as e:
        # フォールバック: 直接Chromeを使用
        print(f"webdriver-manager failed: {e}, trying direct Chrome")
        driver = webdriver.Chrome(options=chrome_options)
    
    return driver

def get_full_page_screenshot(driver):
    """ページ全体のスクリーンショットを取得"""
    # ページの実際の高さを取得
    total_height = driver.execute_script("return document.body.scrollHeight")
    viewport_width = driver.execute_script("return document.body.scrollWidth")
    
    # ウィンドウサイズをページ全体に合わせる
    driver.set_window_size(max(1280, viewport_width), total_height)
    time.sleep(0.5)  # リサイズ後のレンダリング待ち
    
    # スクリーンショット取得
    return driver.get_screenshot_as_png()

def draw_annotations_legacy(screenshot_bytes, elements_data):
    """(旧) スクリーンショットに矢印とIDを描画する（右側のみ）"""
    image = Image.open(io.BytesIO(screenshot_bytes))
    draw = ImageDraw.Draw(image)
    
    # フォント読み込み（失敗したらデフォルト）
    font = None
    font_small = None
    
    if FONT_PATH and os.path.exists(FONT_PATH):
        try:
            font = ImageFont.truetype(FONT_PATH, 26)
            font_small = ImageFont.truetype(FONT_PATH, 22)  # 右側ラベル用（大きめ）
        except Exception as e:
            print(f"フォント読み込みエラー: {e}")
    
    if font is None:
        try:
            font = ImageFont.truetype("Arial", 26)
            font_small = ImageFont.truetype("Arial", 22)
        except:
            font = ImageFont.load_default()
            font_small = ImageFont.load_default()

    # 右側の余白を作るためにカンバスを広げる
    margin_right = 400
    new_width = image.width + margin_right
    new_image = Image.new("RGB", (new_width, image.height), "white")
    new_image.paste(image, (0, 0))
    
    draw = ImageDraw.Draw(new_image)
    
    # 要素をY座標順にソート（上から順番に並ぶように）
    sorted_elements = sorted(elements_data, key=lambda x: x['y'])
    
    # 矢印の色リスト
    colors = [
        "#E60012", "#0066CC", "#009944", "#FF6600",
        "#9933CC", "#00A0E9", "#E4007F", "#8B4513",
    ]
    
    # ラベルの重なりを防ぐためのY座標計算
    label_height = 35
    used_positions = []
    
    def get_non_overlapping_y(target_y):
        candidate_y = max(10, target_y - 12)
        max_attempts = 50
        for _ in range(max_attempts):
            is_overlapping = False
            for pos in used_positions:
                if abs(candidate_y - pos) < label_height:
                    is_overlapping = True
                    candidate_y = pos + label_height
                    break
            if not is_overlapping:
                break
        used_positions.append(candidate_y)
        return candidate_y
    
    def draw_arrow(draw, start, end, color, width=3):
        import math
        draw.line([start, end], fill=color, width=width)
        arrow_size = 12
        angle = math.atan2(end[1] - start[1], end[0] - start[0])
        p1 = end
        p2 = (end[0] - arrow_size * math.cos(angle - math.pi/6),
              end[1] - arrow_size * math.sin(angle - math.pi/6))
        p3 = (end[0] - arrow_size * math.cos(angle + math.pi/6),
              end[1] - arrow_size * math.sin(angle + math.pi/6))
        draw.polygon([p1, p2, p3], fill=color)
    
    circle_numbers = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩',
                      '⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳',
                      '㉑','㉒','㉓','㉔','㉕','㉖','㉗','㉘','㉙','㉚',
                      '㉛','㉜','㉝','㉞','㉟','㊱','㊲','㊳','㊴','㊵',
                      '㊶','㊷','㊸','㊹','㊺','㊻','㊼','㊽','㊾','㊿']

    for i, item in enumerate(sorted_elements):
        color = colors[i % len(colors)]
        target_y = item['y'] + (item['height'] / 2)
        label_x = image.width + 20
        label_y = get_non_overlapping_y(target_y)
        
        display_id = circle_numbers[i] if i < len(circle_numbers) else f"({i + 1})"
        text = f"{display_id}: {item['label'][:12]}" if len(item['label']) > 12 else f"{display_id}: {item['label']}"
        
        try:
            bbox = draw.textbbox((label_x, label_y), text, font=font_small)
            draw.rectangle(bbox, fill="white", outline=color, width=1)
        except:
            pass
        
        draw.text((label_x, label_y), text, fill=color, font=font_small)
        
        arrow_target_x = item['x'] + item['width']
        arrow_target_y = item['y'] + (item['height'] / 2)
        
        start_point = (label_x - 5, label_y + 12)
        end_point = (arrow_target_x + 5, arrow_target_y)
        draw_arrow(draw, start_point, end_point, color, width=3)
        
        draw.rectangle(
            [(item['x'], item['y']), (item['x'] + item['width'], item['y'] + item['height'])],
            outline=color, width=3
        )

    return new_image

def draw_annotations(screenshot_bytes, elements_data):
    """スクリーンショットに矢印とIDを描画する（左右振り分け版）"""
    image = Image.open(io.BytesIO(screenshot_bytes))
    
    # フォント読み込み
    font = None
    font_small = None
    if FONT_PATH and os.path.exists(FONT_PATH):
        try:
            font = ImageFont.truetype(FONT_PATH, 26)
            font_small = ImageFont.truetype(FONT_PATH, 22)
        except:
            pass
    if font is None:
        try:
            font = ImageFont.truetype("Arial", 26)
            font_small = ImageFont.truetype("Arial", 22)
        except:
            font = ImageFont.load_default()
            font_small = ImageFont.load_default()

    # 左右に余白を作る（左400px + 画像 + 右400px）
    margin_side = 400
    new_width = image.width + (margin_side * 2)
    new_image = Image.new("RGB", (new_width, image.height), "white")
    new_image.paste(image, (margin_side, 0)) # 真ん中に画像を配置
    
    draw = ImageDraw.Draw(new_image)
    
    # 要素をY座標順にソート
    sorted_elements = sorted(elements_data, key=lambda x: x['y'])
    
    colors = [
        "#E60012", "#0066CC", "#009944", "#FF6600",
        "#9933CC", "#00A0E9", "#E4007F", "#8B4513",
    ]
    
    # ラベル配置位置の管理（左と右で別管理）
    label_height = 35
    used_positions_left = []
    used_positions_right = []
    
    def get_non_overlapping_y(target_y, is_left):
        """重ならないY座標を取得（左右別）"""
        target_list = used_positions_left if is_left else used_positions_right
        
        candidate_y = max(10, target_y - 12)
        max_attempts = 50
        for _ in range(max_attempts):
            is_overlapping = False
            for pos in target_list:
                if abs(candidate_y - pos) < label_height:
                    is_overlapping = True
                    candidate_y = pos + label_height
                    break
            if not is_overlapping:
                break
        
        target_list.append(candidate_y)
        return candidate_y

    def draw_arrow(draw, start, end, color, width=3):
        import math
        draw.line([start, end], fill=color, width=width)
        arrow_size = 12
        angle = math.atan2(end[1] - start[1], end[0] - start[0])
        p1 = end
        p2 = (end[0] - arrow_size * math.cos(angle - math.pi/6),
              end[1] - arrow_size * math.sin(angle - math.pi/6))
        p3 = (end[0] - arrow_size * math.cos(angle + math.pi/6),
              end[1] - arrow_size * math.sin(angle + math.pi/6))
        draw.polygon([p1, p2, p3], fill=color)

    circle_numbers = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩',
                      '⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳',
                      '㉑','㉒','㉓','㉔','㉕','㉖','㉗','㉘','㉙','㉚',
                      '㉛','㉜','㉝','㉞','㉟','㊱','㊲','㊳','㊴','㊵',
                      '㊶','㊷','㊸','㊹','㊺','㊻','㊼','㊽','㊾','㊿']
    
    # 画面中心（元画像の中心）
    center_x = image.width / 2

    for i, item in enumerate(sorted_elements):
        color = colors[i % len(colors)]
        
        # 元画像の座標系での中心X
        item_center_x = item['x'] + (item['width'] / 2)
        
        # 左右どちらに配置するか判定
        is_left = item_center_x < center_x
        
        # Y座標計算
        item_y_center = item['y'] + (item['height'] / 2)
        label_y = get_non_overlapping_y(item_y_center, is_left)
        
        # ID取得
        display_id = circle_numbers[i] if i < len(circle_numbers) else f"({i + 1})"
        text = f"{display_id}: {item['label'][:12]}" if len(item['label']) > 12 else f"{display_id}: {item['label']}"
        
        # ラベルと矢印のX座標計算
        if is_left:
            # 左側に配置
            label_x = 20 # 左端近く
            
            # 矢印の始点（ラベルの右側）
            # テキスト幅を取得して正確な位置を計算しても良いが、簡易的に固定幅＋余白
            text_width = 250 # 仮の幅
            try:
                bbox = draw.textbbox((0, 0), text, font=font_small)
                text_width = bbox[2] - bbox[0]
            except:
                pass
            
            arrow_start_x = label_x + text_width + 5
            
            # 矢印の終点（要素の左端 + 左マージン分）
            arrow_target_x = item['x'] + margin_side - 5
            
        else:
            # 右側に配置
            label_x = margin_side + image.width + 20
            
            # 矢印の始点（ラベルの左側）
            arrow_start_x = label_x - 5
            
            # 矢印の終点（要素の右端 + 左マージン分）
            arrow_target_x = item['x'] + item['width'] + margin_side + 5

        # ラベル描画
        try:
            bbox = draw.textbbox((label_x, label_y), text, font=font_small)
            draw.rectangle(bbox, fill="white", outline=color, width=1)
        except:
            pass
        draw.text((label_x, label_y), text, fill=color, font=font_small)
        
        # 矢印描画
        start_point = (arrow_start_x, label_y + 12)
        end_point = (arrow_target_x, item_y_center)
        draw_arrow(draw, start_point, end_point, color, width=3)
        
        # 枠線描画（座標は + 左マージン）
        draw.rectangle(
            [(item['x'] + margin_side, item['y']), 
             (item['x'] + item['width'] + margin_side, item['y'] + item['height'])],
            outline=color, width=3
        )

    return new_image

    return new_image

def analyze_html_structure(html_content):
    """HTMLを解析して要素リストとスクリーンショットを返す"""
    
    # 1. 一時ファイルとしてHTMLを保存
    with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as tmp:
        tmp.write(html_content)
        tmp_path = tmp.name

    driver = setup_driver()
    elements_meta = []
    png = None
    
    try:
        # 2. ブラウザで開く
        driver.get(f"file://{tmp_path}")
        time.sleep(1) # レンダリング待ち

        # 3. 解析と座標取得 (JavaScriptで正確な位置を取得)
        # data-labelを持つ要素を探す
        elements = driver.find_elements("css selector", "[data-label]")
        
        # 除外するキーワード（画像/写真関連 - 全セクション共通）
        exclude_keywords_all = ['写真', '画像', 'フォト', 'photo', 'image', 'img', 'ビジュアル', 'MV', '背景']
        
        # パンくずリスト関連の除外キーワード
        exclude_keywords_breadcrumb = ['パンくず', 'breadcrumb', 'topicpath', 'pankuzu']
        
        # CTA関連の除外キーワード
        exclude_keywords_cta = ['cta', 'contact', 'reservation', 'button', 'btn', 'お問い合わせ', '資料請求', '申し込み', 'CV', 'action']
        
        # ヒーローセクションのみ除外するキーワード
        exclude_keywords_hero = ['大見出し', 'サブタイトル', 'タイトル', '見出し英語', '見出しEN', '見出し']
        
        for idx, elem in enumerate(elements):
            # 表示されていない要素（titleなど）は座標取得でエラーになるため除外するかチェック
            if not elem.is_displayed():
                continue

            # data属性から情報取得
            section = elem.get_attribute("data-section") or ""  # セクション名
            label = elem.get_attribute("data-label") or ""  # 要素名
            limit = elem.get_attribute("data-limit") or ""  # 文字数制限
            text = elem.text.strip()
            
            # 写真・画像関連は全セクションで除外
            if any(keyword.lower() in label.lower() for keyword in exclude_keywords_all):
                continue
            
            # パンくずリストは除外（セクション名またはラベル名にキーワードが含まれる場合）
            if any(keyword.lower() in label.lower() for keyword in exclude_keywords_breadcrumb) or \
               any(keyword.lower() in section.lower() for keyword in exclude_keywords_breadcrumb):
                continue
            
            # CTA関連は除外（セクション名またはラベル名にキーワードが含まれる場合）
            if any(keyword.lower() in label.lower() for keyword in exclude_keywords_cta) or \
               any(keyword.lower() in section.lower() for keyword in exclude_keywords_cta):
                continue
            
            # ヒーローセクションの見出し関連は除外
            if 'ヒーロー' in section.lower() or 'hero' in section.lower():
                if any(keyword.lower() in label.lower() for keyword in exclude_keywords_hero):
                    continue
            
            # 座標取得
            rect = elem.rect # x, y, width, height
            
            # リストに追加
            elements_meta.append({
                "section": section,
                "label": label,
                "text": text,
                "limit": limit,
                "x": rect['x'],
                "y": rect['y'],
                "width": rect['width'],
                "height": rect['height']
            })
        
        # Y座標でソート（上から順番に）
        elements_meta.sort(key=lambda x: x['y'])
        
        # 4. スクリーンショット撮影（ページ全体）
        png = get_full_page_screenshot(driver)
        
    finally:
        driver.quit()
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
            
    return elements_meta, png

def create_excel_file(selected_elements, original_screenshot_bytes):
    """選択された要素に基づきExcelと注釈付き画像を生成する"""
    
    # 丸数字のリスト
    circle_numbers = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩',
                      '⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳',
                      '㉑','㉒','㉓','㉔','㉕','㉖','㉗','㉘','㉙','㉚',
                      '㉛','㉜','㉝','㉞','㉟','㊱','㊲','㊳','㊴','㊵',
                      '㊶','㊷','㊸','㊹','㊺','㊻','㊼','㊽','㊾','㊿']
    
    data_rows = []
    processed_elements = [] # 画像描画用（ID付き）
    
    # IDの割り当て（選択された要素のみ連番）
    for i, item in enumerate(selected_elements):
        if i < len(circle_numbers):
            row_id = circle_numbers[i]
        else:
            row_id = f"({i + 1})"
        
        # 描画用にIDを追加した辞書を作成
        item_with_id = item.copy()
        item_with_id['id'] = row_id
        processed_elements.append(item_with_id)
        
        # Excelデータに追加
        data_rows.append({
            "ID": row_id,
            "セクション": item['section'],
            "要素": item['label'],
            "ワイヤー記載（参考）": item['text'],
            "クライアント入力": "",
            "文字数目安": item['limit'],
            "現在文字数": ""
        })

    # 画像加工（矢印描画）
    annotated_img = draw_annotations(original_screenshot_bytes, processed_elements)
    
    # Excel生成
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: リスト
        df = pd.DataFrame(data_rows)
        df.to_excel(writer, sheet_name=SHEET1_NAME, index=False)
        
        # Sheet 1の装飾
        worksheet1 = writer.sheets[SHEET1_NAME]
        
        # 列幅設定
        worksheet1.column_dimensions['A'].width = 12
        worksheet1.column_dimensions['B'].width = 16
        worksheet1.column_dimensions['C'].width = 16
        worksheet1.column_dimensions['D'].width = 45
        worksheet1.column_dimensions['E'].width = 45
        worksheet1.column_dimensions['F'].width = 10
        worksheet1.column_dimensions['G'].width = 10
        
        # スタイル定義
        header_fill = PatternFill(start_color='4A7C59', end_color='4A7C59', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        input_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
        input_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        normal_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
        
        # ヘッダー行スタイル
        for cell in worksheet1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # データ行スタイル
        for row_idx, row in enumerate(worksheet1.iter_rows(min_row=2, max_row=worksheet1.max_row), start=2):
            worksheet1.row_dimensions[row_idx].height = 50
            for cell in row:
                cell.alignment = normal_alignment
                cell.border = thin_border
                if cell.column_letter == 'E':
                    cell.fill = input_fill
                    cell.alignment = input_alignment
                if cell.column_letter == 'G':
                    cell.value = f'=LEN(E{row_idx})'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet1.row_dimensions[1].height = 30
        worksheet1.freeze_panes = 'A2'
        
        # Sheet 2: 画像貼り付け
        pd.DataFrame(["以下画像参照"]).to_excel(writer, sheet_name=SHEET2_NAME, index=False, header=False)
        worksheet2 = writer.sheets[SHEET2_NAME]
        
        img_byte_arr = io.BytesIO()
        annotated_img.save(img_byte_arr, format='PNG')
        img_to_excel = openpyxl_image(img_byte_arr)
        worksheet2.add_image(img_to_excel, 'A1')

    output.seek(0)
    return output

# OpenPyXLのスタイル関連インポート
from openpyxl.drawing.image import Image as openpyxl_image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==========================================
# UI構築 (Streamlit)
# ==========================================
st.set_page_config(page_title="ワイヤー原稿ツール", layout="wide")

st.title("📑 ワイヤーフレーム原稿依頼書ジェネレーター")
st.markdown("""
HTMLファイルをアップロードし、必要な項目を選択してExcel原稿を作成します。
""")

# セッション状態の初期化
if 'step' not in st.session_state:
    st.session_state['step'] = 'upload'
if 'analyzed_data' not in st.session_state:
    st.session_state['analyzed_data'] = []
if 'screenshot' not in st.session_state:
    st.session_state['screenshot'] = None

# ステップ1: ファイルアップロード
if st.session_state['step'] == 'upload':
    uploaded_file = st.file_uploader("HTMLファイルをドラッグ＆ドロップ", type=["html", "htm"])

    if uploaded_file is not None:
        if st.button("ファイルを解析する", type="primary"):
            with st.spinner("ファイルを解析中... ブラウザレンダリングを実行しています"):
                try:
                    html_bytes = uploaded_file.read()
                    
                    # HTML解析実行
                    elements_meta, png_bytes = analyze_html_structure(html_bytes)
                    
                    # セッションに保存
                    st.session_state['analyzed_data'] = elements_meta
                    st.session_state['screenshot'] = png_bytes
                    st.session_state['filename'] = uploaded_file.name
                    st.session_state['step'] = 'preview'
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"解析中にエラーが発生しました: {e}")

# ステップ2: プレビューと選択
elif st.session_state['step'] == 'preview':
    st.success("解析完了！ 出力する項目を選択してください。（チェックを変更すると画像が更新されます）")
    
    # 画面分割（左：リスト、右：プレビュー画像）
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("要素リスト")
        
        # データフレーム作成（チェックボックス用）
        df_preview = pd.DataFrame(st.session_state['analyzed_data'])
        
        selected_elements = []  # 初期化
        
        # 表示用カラムの整理
        if not df_preview.empty:
            # 選択用カラムを先頭に追加（デフォルトTrue）
            if "選択" not in df_preview.columns:
                df_preview.insert(0, "選択", True)
            
            # ユーザーが編集可能なデータエディタを表示
            edited_df = st.data_editor(
                df_preview[['選択', 'section', 'label', 'text', 'limit']],
                column_config={
                    "選択": st.column_config.CheckboxColumn(
                        "出力",
                        help="チェックを外すとExcelに出力されません",
                        default=True,
                    ),
                    "section": "セクション",
                    "label": "要素名",
                    "text": "テキスト内容",
                    "limit": "文字数制限"
                },
                disabled=["section", "label", "text", "limit"],
                hide_index=True,
                height=600,
                key="data_editor" # キーを指定して状態を管理
            )
            
            # 選択された行のみを抽出
            selected_indices = edited_df[edited_df['選択'] == True].index
            selected_elements = [st.session_state['analyzed_data'][i] for i in selected_indices]
            
            st.info(f"全 {len(st.session_state['analyzed_data'])} 項目中、 {len(selected_elements)} 項目を選択中")
            
        else:
            st.warning("有効な要素が見つかりませんでした。")
            selected_elements = []

        st.divider()
        
        # Excel生成ボタン（左カラム下に配置）
        if st.button("Excelファイルを生成する", type="primary", disabled=len(selected_elements)==0):
            with st.spinner("Excelを作成中..."):
                try:
                    excel_file = create_excel_file(selected_elements, st.session_state['screenshot'])
                    
                    # 生成完了アニメーション
                    st.balloons()
                    
                    # ファイル名生成
                    original_name = st.session_state.get('filename', 'output.html')
                    base_name = original_name.rsplit('.', 1)[0]
                    excel_filename = f"{base_name}.xlsx"
                    
                    st.download_button(
                        label=f"📥 {excel_filename} をダウンロード",
                        data=excel_file,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except Exception as e:
                    st.error(f"生成エラー: {e}")
        
        if st.button("最初に戻る"):
            st.session_state['step'] = 'upload'
            st.session_state['analyzed_data'] = []
            st.session_state['screenshot'] = None
            st.rerun()

    with col2:
        st.subheader("プレビュー")
        
        if st.session_state['screenshot'] is not None:
            # 選択された要素に基づいて画像をリアルタイム生成
            
            # ID割り当て用（処理用にコピー）
            processed_elements_preview = []
            
            # 丸数字のリスト
            circle_numbers = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩',
                              '⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳',
                              '㉑','㉒','㉓','㉔','㉕','㉖','㉗','㉘','㉙','㉚',
                              '㉛','㉜','㉝','㉞','㉟','㊱','㊲','㊳','㊴','㊵',
                              '㊶','㊷','㊸','㊹','㊺','㊻','㊼','㊽','㊾','㊿']

            # 選択された要素にIDを振る
            for i, item in enumerate(selected_elements):
                if i < len(circle_numbers):
                    row_id = circle_numbers[i]
                else:
                    row_id = f"({i + 1})"
                
                item_with_id = item.copy()
                item_with_id['id'] = row_id
                processed_elements_preview.append(item_with_id)
            
            # 画像描画
            preview_img = draw_annotations(st.session_state['screenshot'], processed_elements_preview)
            
            st.image(preview_img, caption="選択項目のワイヤーフレーム", use_container_width=True)
        else:
            st.write("画像がありません")